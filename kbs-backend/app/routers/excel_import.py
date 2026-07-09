from typing import Any, Dict, List
import io
import re

from fastapi import APIRouter, Depends, HTTPException, Header, UploadFile, File
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import desc

from app.database import get_db
from app.models import Package, ImportHistory
from app.routers.cms import get_admin_role
import openpyxl

router = APIRouter(prefix="/api/cms/excel", tags=["excel_import"])

REQUIRED_COLUMNS = ["title", "destination", "price", "duration"]

COLUMN_MAPPING = {
    "title": ["title", "package title", "packagetitle"],
    "destination": ["destination", "location", "dest"],
    "price": ["price", "cost", "rate"],
    "duration": ["duration", "days", "time", "period"],
    "image": ["image", "image url", "imageurl", "photo", "pic"],
    "rating": ["rating", "score"],
    "rating_label": ["rating label", "ratinglabel", "rating_label", "badge"],
    "reviews": ["reviews", "num reviews", "review count", "reviews_count"],
    "stars": ["stars", "star rating", "star_rating"],
    "bullets": ["bullets", "highlights summary", "features"],
    "tags": ["tags", "amenities"],
    "categories": ["categories", "category", "theme"],
    "tour_types": ["tour types", "tourtype", "tour_types", "tourtypes"],
    "popular_filters": ["popular filters", "popularfilters", "popular_filters", "filters"],
    "highlights": ["highlights", "key highlights", "itinerary highlights"],
    "inclusions": ["inclusions", "included"],
    "exclusions": ["exclusions", "excluded"],
    "stay_transfers": ["stay transfers", "staytransfers", "stay_transfers", "stay and transfers", "stay & transfers"],
    "slug": ["slug", "url slug", "url_slug"]
}

def clean_header(val: Any) -> str:
    if not val:
        return ""
    return str(val).strip().lower()

def parse_array_field(val: Any) -> List[str]:
    if not val:
        return []
    val_str = str(val).strip()
    if not val_str:
        return []
    # Split by semicolon, comma, or newline
    parts = re.split(r'[;\n\r,]+', val_str)
    return [p.strip() for p in parts if p.strip()]

def slugify(title: str) -> str:
    s = title.lower().strip()
    s = re.sub(r'[^a-z0-9\s-]', '', s)
    s = re.sub(r'[\s-]+', '-', s)
    return s

@router.post("/validate")
def validate_excel(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    role: str = Depends(get_admin_role)
):
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="Invalid file type. Only Excel files (.xlsx, .xls) are allowed.")

    try:
        contents = file.file.read()
        wb = openpyxl.load_workbook(io.BytesIO(contents), data_only=True)
        sheet = wb.active
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Failed to read Excel file: {str(e)}")

    # Extract headers
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        raise HTTPException(status_code=400, detail="Excel sheet is empty.")

    header_row = rows[0]
    headers_dict = {}
    
    # Map headers to canonical fields
    for idx, cell in enumerate(header_row):
        if cell is None:
            continue
        cleaned = clean_header(cell)
        matched = False
        for field, aliases in COLUMN_MAPPING.items():
            if cleaned in aliases:
                headers_dict[field] = idx
                matched = True
                break
        if not matched:
            headers_dict[cleaned] = idx

    # Check required columns
    missing_required = [col for col in REQUIRED_COLUMNS if col not in headers_dict]
    if missing_required:
        return {
            "valid": False,
            "errors": [f"Missing required columns: {', '.join(missing_required)}. Please ensure headers align with: Title, Destination, Price, Duration."],
            "total_rows": 0,
            "preview_rows": [],
            "failed_rows_json": []
        }

    preview_rows = []
    errors = []
    failed_rows = []
    valid_rows_count = 0

    # Parse rows
    for r_idx, row in enumerate(rows[1:], start=2):
        # Skip completely empty rows
        if not any(cell is not None for cell in row):
            continue

        row_errors = []

        # Helper to get cell value securely
        def get_val(field: str, default: Any = None) -> Any:
            if field in headers_dict:
                idx = headers_dict[field]
                if idx < len(row):
                    return row[idx]
            return default

        # Required fields check
        title = get_val("title")
        destination = get_val("destination")
        price_val = get_val("price")
        duration = get_val("duration")

        if not title:
            row_errors.append("Title is required.")
        if not destination:
            row_errors.append("Destination is required.")
        
        price = 0
        if price_val is None:
            row_errors.append("Price is required.")
        else:
            try:
                price = int(float(str(price_val).replace('$', '').replace(',', '').strip()))
                if price < 0:
                    row_errors.append("Price cannot be negative.")
            except ValueError:
                row_errors.append(f"Price must be a valid number (got '{price_val}').")

        if not duration:
            row_errors.append("Duration is required (e.g. 5D/4N).")

        # Parse optional fields
        slug = get_val("slug")
        if slug:
            slug = slugify(str(slug))
        elif title:
            slug = slugify(str(title))

        image = get_val("image", "/src/assets/hero.webp")
        rating_val = get_val("rating", 5.0)
        reviews_val = get_val("reviews", 5)
        stars_val = get_val("stars", 5)

        rating = 5.0
        try:
            rating = float(rating_val)
        except ValueError:
            row_errors.append(f"Rating must be a number (got '{rating_val}').")

        reviews = 5
        try:
            reviews = int(reviews_val)
        except ValueError:
            row_errors.append(f"Reviews must be an integer (got '{reviews_val}').")

        stars = 5
        try:
            stars = int(stars_val)
            if stars < 1 or stars > 5:
                row_errors.append(f"Stars must be between 1 and 5 (got '{stars_val}').")
        except ValueError:
            row_errors.append(f"Stars must be an integer (got '{stars_val}').")

        rating_label = get_val("rating_label")
        if not rating_label:
            if rating >= 4.5:
                rating_label = "Superb"
            elif rating >= 4.0:
                rating_label = "Excellent"
            else:
                rating_label = "Good"

        bullets = parse_array_field(get_val("bullets"))
        tags = parse_array_field(get_val("tags"))
        categories = parse_array_field(get_val("categories"))
        tour_types = parse_array_field(get_val("tour_types"))
        popular_filters = parse_array_field(get_val("popular_filters"))
        highlights = parse_array_field(get_val("highlights"))
        inclusions = parse_array_field(get_val("inclusions"))
        exclusions = parse_array_field(get_val("exclusions"))
        stay_transfers = get_val("stay_transfers", "")

        # Default tags/categories if empty
        if not categories:
            categories = ["Weekend"]
        if not tour_types:
            tour_types = ["Family Tour"]

        parsed_data = {
            "title": str(title) if title else "",
            "destination": str(destination) if destination else "",
            "price": price,
            "duration": str(duration) if duration else "",
            "slug": slug,
            "image": str(image),
            "rating": rating,
            "ratingLabel": rating_label,
            "reviews": reviews,
            "stars": stars,
            "bullets": bullets,
            "tags": tags,
            "categories": categories,
            "tourTypes": tour_types,
            "popularFilters": popular_filters,
            "highlights": highlights,
            "inclusions": inclusions,
            "exclusions": exclusions,
            "stayTransfers": str(stay_transfers)
        }

        # Check action type (insert vs update)
        action = "insert"
        if slug:
            existing = db.query(Package).filter(Package.slug == slug, Package.is_active == True).first()
            if existing:
                action = "update"
                parsed_data["id"] = existing.id

        if row_errors:
            failed_rows.append({
                "row_number": r_idx,
                "data": {str(header_row[i]): row[i] for i in range(len(row)) if i < len(header_row)},
                "errors": row_errors
            })
            errors.append(f"Row {r_idx}: {'; '.join(row_errors)}")
        else:
            valid_rows_count += 1
            parsed_data["rowNumber"] = r_idx
            parsed_data["action"] = action
            preview_rows.append(parsed_data)

    return {
        "valid": len(failed_rows) == 0,
        "total_rows": len(rows) - 1,
        "preview_rows": preview_rows,
        "failed_rows_json": failed_rows,
        "errors": errors
    }

@router.post("/confirm")
def confirm_excel_import(
    payload: Dict[str, Any],
    db: Session = Depends(get_db),
    role: str = Depends(get_admin_role),
    editor_name: str = Header("Admin", alias="X-Editor-Name")
):
    preview_rows = payload.get("preview_rows", [])
    failed_rows = payload.get("failed_rows", [])
    file_name = payload.get("file_name", "uploaded_packages.xlsx")

    imported_count = 0
    updated_count = 0
    failed_count = len(failed_rows)
    total_rows = len(preview_rows) + failed_count

    logs = []

    for row in preview_rows:
        try:
            slug = row.get("slug")
            # Ensure unique slug for insertions
            if row.get("action") == "insert":
                base_slug = slug
                counter = 1
                while db.query(Package).filter(Package.slug == slug).first():
                    slug = f"{base_slug}-{counter}"
                    counter += 1
                
                pkg = Package(
                    slug=slug,
                    title=row["title"],
                    destination=row["destination"],
                    price=row["price"],
                    duration=row["duration"],
                    image=row["image"],
                    rating=row["rating"],
                    rating_label=row["ratingLabel"],
                    reviews=row["reviews"],
                    stars=row["stars"],
                    bullets=row["bullets"],
                    tags=row["tags"],
                    categories=row["categories"],
                    tour_types=row["tourTypes"],
                    popular_filters=row["popularFilters"],
                    highlights=row["highlights"],
                    inclusions=row["inclusions"],
                    exclusions=row["exclusions"],
                    stay_transfers=row["stayTransfers"],
                    is_published=True,
                    is_active=True
                )
                db.add(pkg)
                imported_count += 1
                logs.append(f"Row {row['rowNumber']}: Imported new package '{row['title']}' (slug: {slug})")
            else:
                # Update
                pkg_id = row.get("id")
                pkg = db.query(Package).filter(Package.id == pkg_id).first()
                if pkg:
                    pkg.title = row["title"]
                    pkg.destination = row["destination"]
                    pkg.price = row["price"]
                    pkg.duration = row["duration"]
                    pkg.image = row["image"]
                    pkg.rating = row["rating"]
                    pkg.rating_label = row["ratingLabel"]
                    pkg.reviews = row["reviews"]
                    pkg.stars = row["stars"]
                    pkg.bullets = row["bullets"]
                    pkg.tags = row["tags"]
                    pkg.categories = row["categories"]
                    pkg.tour_types = row["tourTypes"]
                    pkg.popular_filters = row["popularFilters"]
                    pkg.highlights = row["highlights"]
                    pkg.inclusions = row["inclusions"]
                    pkg.exclusions = row["exclusions"]
                    pkg.stay_transfers = row["stayTransfers"]
                    pkg.is_published = True
                    updated_count += 1
                    logs.append(f"Row {row['rowNumber']}: Updated package '{row['title']}' (ID: {pkg_id})")
                else:
                    failed_count += 1
                    logs.append(f"Row {row['rowNumber']}: Package update target not found (ID: {pkg_id})")

        except Exception as ex:
            failed_count += 1
            logs.append(f"Row {row.get('rowNumber', 'unknown')}: Failed to save due to backend error: {str(ex)}")

    # Add failed rows description to logs
    for f in failed_rows:
        logs.append(f"Row {f['row_number']}: Failed validation. Errors: {', '.join(f['errors'])}")

    # Record to ImportHistory
    status = "success"
    if failed_count > 0:
        status = "partial_success" if (imported_count > 0 or updated_count > 0) else "failed"

    history_entry = ImportHistory(
        file_name=file_name,
        uploaded_by=editor_name,
        total_rows=total_rows,
        imported=imported_count,
        updated=updated_count,
        failed=failed_count,
        status=status,
        logs=logs,
        failed_rows_json=failed_rows
    )
    db.add(history_entry)
    db.commit()

    return {
        "status": status,
        "imported": imported_count,
        "updated": updated_count,
        "failed": failed_count,
        "logs": logs
    }

@router.get("/history")
def get_import_history(
    db: Session = Depends(get_db),
    role: str = Depends(get_admin_role)
):
    return db.query(ImportHistory).order_by(desc(ImportHistory.created_at)).all()

@router.get("/history/{import_id}/failed-excel")
def download_failed_excel(
    import_id: int,
    db: Session = Depends(get_db),
    role: str = Depends(get_admin_role)
):
    entry = db.query(ImportHistory).filter(ImportHistory.id == import_id).first()
    if not entry:
        raise HTTPException(status_code=404, detail="Import record not found")

    failed_rows = entry.failed_rows_json or []
    if not failed_rows:
        raise HTTPException(status_code=400, detail="No failed rows associated with this import.")

    # Create a new workbook
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = "Failed Rows"

    # Extract headers from the first failed row data keys
    headers = []
    if failed_rows:
        headers = list(failed_rows[0]["data"].keys())
    
    # Append validation errors column at the end
    headers_with_errors = headers + ["Validation Errors"]
    sheet.append(headers_with_errors)

    for row_item in failed_rows:
        row_data = row_item["data"]
        row_errors = "; ".join(row_item["errors"])
        
        row_values = []
        for h in headers:
            row_values.append(row_data.get(h, ""))
        row_values.append(row_errors)
        sheet.append(row_values)

    # Save to dynamic streaming stream
    out = io.BytesIO()
    wb.save(out)
    out.seek(0)

    filename = f"failed_rows_import_{import_id}.xlsx"
    return StreamingResponse(
        out,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )
